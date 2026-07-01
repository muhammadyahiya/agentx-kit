"""Document loaders for agentx RAG pipeline.

Supports: PDF, Excel (.xlsx/.xls), CSV, Word (.docx), plain text, Markdown.
Each loader returns a list of plain-text strings ready for chunking.

Install the matching extras:
    agentx-kit[rag]         — langchain-community, text-splitters, chromadb
    agentx-kit[pdf]         — pypdf
    agentx-kit[excel]       — openpyxl
"""
from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Literal

from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

DocumentType = Literal["pdf", "excel", "csv", "word", "txt", "md", "auto"]


# ──────────────────────────────────────────────────────────────────────────────
# Pydantic config
# ──────────────────────────────────────────────────────────────────────────────

class LoaderConfig(BaseModel):
    """Configuration for document loading."""

    doc_type: DocumentType = "auto"
    encoding: str = "utf-8"
    # PDF options
    extract_images: bool = False
    # Excel/CSV options
    sheet_name: str | int | None = None
    max_rows: int | None = Field(default=None, description="Max rows to load from Excel/CSV")
    # Text joining
    page_separator: str = "\n\n--- page break ---\n\n"

    model_config = {"extra": "allow"}


# ──────────────────────────────────────────────────────────────────────────────
# Extension → type mapping
# ──────────────────────────────────────────────────────────────────────────────

_EXT_MAP: dict[str, DocumentType] = {
    ".pdf": "pdf",
    ".xlsx": "excel",
    ".xls": "excel",
    ".xlsm": "excel",
    ".csv": "csv",
    ".docx": "word",
    ".doc": "word",
    ".txt": "txt",
    ".md": "md",
    ".markdown": "md",
    ".rst": "txt",
}


def _detect_type(path: Path) -> DocumentType:
    return _EXT_MAP.get(path.suffix.lower(), "txt")


# ──────────────────────────────────────────────────────────────────────────────
# Per-format loaders
# ──────────────────────────────────────────────────────────────────────────────

def _load_pdf(path: Path, cfg: LoaderConfig) -> list[str]:
    """Load PDF — tries pypdf first, then langchain PyPDFLoader."""
    try:
        import pypdf  # type: ignore

        pages: list[str] = []
        with pypdf.PdfReader(str(path)) as reader:
            for page in reader.pages:
                text = page.extract_text() or ""
                if text.strip():
                    pages.append(text)
        logger.info("Loaded PDF via pypdf: %s (%d pages)", path.name, len(pages))
        return pages
    except ImportError:
        pass

    try:
        from langchain_community.document_loaders import PyPDFLoader  # type: ignore

        loader = PyPDFLoader(str(path), extract_images=cfg.extract_images)
        docs = loader.load()
        pages = [d.page_content for d in docs if d.page_content.strip()]
        logger.info("Loaded PDF via langchain PyPDFLoader: %s (%d pages)", path.name, len(pages))
        return pages
    except ImportError:
        pass

    raise ImportError(
        f"Cannot load PDF '{path.name}'. Install a PDF reader:\n"
        "    pip install pypdf\n"
        "    # or: pip install 'agentx-kit[pdf]'"
    )


def _load_excel(path: Path, cfg: LoaderConfig) -> list[str]:
    """Load Excel — returns one text block per sheet (or selected sheet)."""
    try:
        import openpyxl  # type: ignore

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheet_names = (
            [cfg.sheet_name] if cfg.sheet_name is not None else wb.sheetnames
        )
        pages: list[str] = []
        for name in sheet_names:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if cfg.max_rows:
                rows = rows[: cfg.max_rows]
            header = rows[0] if rows else []
            lines = ["\t".join(str(c) if c is not None else "" for c in header)]
            lines += [
                "\t".join(str(c) if c is not None else "" for c in row)
                for row in rows[1:]
            ]
            text = f"Sheet: {name}\n" + "\n".join(lines)
            pages.append(text)
        logger.info(
            "Loaded Excel via openpyxl: %s (%d sheets)", path.name, len(pages)
        )
        return pages
    except ImportError:
        pass

    try:
        import pandas as pd  # type: ignore

        xl = pd.ExcelFile(str(path))
        sheet_names = (
            [cfg.sheet_name] if cfg.sheet_name is not None else xl.sheet_names
        )
        pages = []
        for name in sheet_names:
            df = xl.parse(name, nrows=cfg.max_rows)
            pages.append(f"Sheet: {name}\n{df.to_string(index=False)}")
        logger.info("Loaded Excel via pandas: %s (%d sheets)", path.name, len(pages))
        return pages
    except ImportError:
        pass

    raise ImportError(
        f"Cannot load Excel '{path.name}'. Install openpyxl:\n"
        "    pip install openpyxl\n"
        "    # or: pip install 'agentx-kit[excel]'"
    )


def _load_csv(path: Path, cfg: LoaderConfig) -> list[str]:
    """Load CSV as a formatted text block."""
    lines: list[str] = []
    with path.open(encoding=cfg.encoding, newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if cfg.max_rows and i >= cfg.max_rows:
                break
            lines.append("\t".join(row))
    logger.info("Loaded CSV: %s (%d rows)", path.name, len(lines))
    return ["\n".join(lines)]


def _load_word(path: Path, cfg: LoaderConfig) -> list[str]:
    """Load .docx via python-docx or langchain."""
    try:
        import docx  # type: ignore

        doc = docx.Document(str(path))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        logger.info("Loaded Word via python-docx: %s", path.name)
        return [text] if text else []
    except ImportError:
        pass

    try:
        from langchain_community.document_loaders import Docx2txtLoader  # type: ignore

        loader = Docx2txtLoader(str(path))
        docs = loader.load()
        pages = [d.page_content for d in docs if d.page_content.strip()]
        logger.info("Loaded Word via langchain Docx2txtLoader: %s", path.name)
        return pages
    except ImportError:
        pass

    raise ImportError(
        f"Cannot load Word doc '{path.name}'. Install python-docx:\n"
        "    pip install python-docx"
    )


def _load_text(path: Path, cfg: LoaderConfig) -> list[str]:
    text = path.read_text(encoding=cfg.encoding)
    logger.debug("Loaded text file: %s (%d chars)", path.name, len(text))
    return [text] if text.strip() else []


# ──────────────────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────────────────

def load_document(path: str | Path, config: LoaderConfig | None = None) -> list[str]:
    """Load a single document and return its content as a list of text strings.

    Each element in the returned list is a logical "page" or "section" —
    pass the result directly to ``build_index_from_texts()``.

    Args:
        path: Path to the document (PDF, Excel, CSV, Word, TXT, MD).
        config: Optional ``LoaderConfig`` to control loader behaviour.

    Returns:
        List of text strings (one per page / sheet / section).
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Document not found: {p}")

    cfg = config or LoaderConfig()
    doc_type = cfg.doc_type if cfg.doc_type != "auto" else _detect_type(p)

    dispatch = {
        "pdf": _load_pdf,
        "excel": _load_excel,
        "csv": _load_csv,
        "word": _load_word,
        "txt": _load_text,
        "md": _load_text,
    }
    loader_fn = dispatch.get(doc_type, _load_text)
    return loader_fn(p, cfg)


def load_directory(
    directory: str | Path,
    config: LoaderConfig | None = None,
    glob: str = "**/*",
    skip_errors: bool = True,
) -> dict[str, list[str]]:
    """Load all supported documents from a directory.

    Args:
        directory: Path to scan.
        config: Shared loader config applied to every document.
        glob: Glob pattern to filter files (default: all files recursively).
        skip_errors: Log and skip files that fail to load (True) or raise (False).

    Returns:
        ``{filename: [page_text, ...]}`` dict.
    """
    d = Path(directory)
    supported_exts = set(_EXT_MAP.keys())
    results: dict[str, list[str]] = {}

    for fp in sorted(d.glob(glob)):
        if not fp.is_file():
            continue
        if fp.suffix.lower() not in supported_exts:
            continue
        if fp.name.startswith(".") or fp.name == "README.md":
            continue
        try:
            texts = load_document(fp, config)
            if texts:
                results[fp.name] = texts
                logger.debug("Loaded '%s': %d sections", fp.name, len(texts))
        except Exception as exc:  # noqa: BLE001
            if skip_errors:
                logger.warning("Skipping '%s': %s", fp.name, exc)
            else:
                raise

    logger.info(
        "Loaded %d document(s) from %s", len(results), d
    )
    return results
